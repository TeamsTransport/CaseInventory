import os
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

INVALID_SHEET_CHARS = r'[\[\]\*\/\?:\\]'  # Excel-invalid characters for sheet names

def make_unique_title(base_title: str, existing: set) -> str:
    """
    Ensure the title is unique against `existing`. If needed,
    append ' (2)', ' (3)', ... while keeping length <= 31.
    """
    title = base_title or "Sheet"
    if title not in existing:
        existing.add(title)
        return title

    i = 2
    while True:
        suffix = f" ({i})"
        candidate = (base_title[: 31 - len(suffix)]) + suffix
        if candidate not in existing:
            existing.add(candidate)
            return candidate
        i += 1

def sanitize_sheet_title(name, existing_titles: set) -> str:
    """
    Replace invalid characters, strip leading/trailing apostrophes,
    trim to 31 characters, and make unique.
    """
    s = str(name).strip()
    # Replace invalid characters with '-'
    s = re.sub(INVALID_SHEET_CHARS, "-", s)
    # Excel dislikes leading/trailing apostrophes in sheet names
    s = s.strip("'")
    # Trim to 31 characters (Excel limit)
    s = s[:31] if s else "Sheet"
    return make_unique_title(s, existing_titles)

# Load the latest generated Excel file
files = [f for f in os.listdir() if f.startswith("LoblawsConsolidatedInventory_") and f.endswith(".xlsx")]
latest_file = max(files, key=os.path.getctime)
wb = load_workbook(latest_file)
ws_master = wb["Master"]

# Insert 5 rows at the top
ws_master.insert_rows(1, amount=5)

# Add TEAMS logo image from same folder
logo_path = "TeamsLogo.png"
if os.path.exists(logo_path):
    img = Image(logo_path)
    img.anchor = "B2"
    ws_master.add_image(img)

# Get previous month info
today = datetime.today()
first_day_this_month = datetime(today.year, today.month, 1)
last_day_prev_month = first_day_this_month - timedelta(days=1)
prev_month_name = last_day_prev_month.strftime("%B")
prev_month_day = last_day_prev_month.day
current_year = today.year

# Merge and format E2:G2
ws_master.merge_cells("E2:G2")
ws_master["E2"].value = f"Inventory {prev_month_name} {prev_month_day} {current_year}"
ws_master["E2"].alignment = Alignment(horizontal="center", vertical="center")

# Merge and format E3:G3
ws_master.merge_cells("E3:G3")
ws_master["E3"].value = "SAP 4021445"
ws_master["E3"].alignment = Alignment(horizontal="center", vertical="center")

# Define header row and formatting
header_row = 6
headers = [
    'Store', 'PM', 'PO', 'Shipper Order #', 'Line Up #', 'Case #', 'Case Model #', 'Serial #',
    'Arrived @ Warehouse', 'Storage Starts', 'Storage Ends', 'Scheduled Date','Scheduled Time', 'Warehouse Location', 
    'Damage Y-N', 'Delivery Trailer', '# Days In Storage', 'Square Footage', 'Storage Charge', 'Extended Price'
]

for col_num, header in enumerate(headers, start=1):
    cell = ws_master.cell(row=header_row, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

# Set row height
ws_master.row_dimensions[header_row].height = 39

# Apply thin borders to all cells from row 6 onward, columns A to T
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws_master.iter_rows(min_row=6, max_row=ws_master.max_row, min_col=1, max_col=20):
    for cell in row:
        cell.border = thin_border

# Freeze top 6 rows
ws_master.freeze_panes = "A7"

# Format date columns to MMDDYY
date_columns = ['Arrived @ Warehouse', 'Storage Starts', 'Storage Ends', 'Scheduled Date']
for row in ws_master.iter_rows(min_row=7, max_row=ws_master.max_row):
    for col_name in date_columns:
        if col_name in headers:
            col_index = headers.index(col_name)
            cell = row[col_index]
            cell.number_format = "MM-DD-YY"

# Format currency columns
currency_columns = ['Storage Charge', 'Extended Price']
for row in ws_master.iter_rows(min_row=7, max_row=ws_master.max_row):
    for col_name in currency_columns:
        if col_name in headers:
            col_index = headers.index(col_name)
            cell = row[col_index]
            cell.number_format = '"$"#,##0.00'

# List of columns to check
columns_to_clear = ['Scheduled Time']

# Clear cells with 'NaN' or 'None'
for row in ws_master.iter_rows(min_row=7, max_row=ws_master.max_row):
    for col_name in columns_to_clear:
        if col_name in headers:
            col_index = headers.index(col_name)
            cell = row[col_index]
            if str(cell.value).strip().lower() in ['nan', 'none']:
                cell.value = None

# Adjust formulas to start from row 7
header = [cell.value for cell in ws_master[header_row]]
col_indices = {col: idx + 1 for idx, col in enumerate(header)}
max_row = ws_master.max_row
for row in range(header_row + 1, max_row + 1):
    try:
        start_col = get_column_letter(col_indices['Storage Starts'])
        end_col = get_column_letter(col_indices['Storage Ends'])
        days_col = get_column_letter(col_indices['# Days In Storage'])
        charge_col = get_column_letter(col_indices['Storage Charge'])
        sqft_col = get_column_letter(col_indices['Square Footage'])
        extended_col = get_column_letter(col_indices['Extended Price'])
        ws_master[f"{days_col}{row}"] = f"=({end_col}{row}-{start_col}{row})+1"
        ws_master[f"{extended_col}{row}"] = f"=((( {charge_col}{row} * {sqft_col}{row} ) / 30) * {days_col}{row})+60"
    except KeyError:
        continue

# Create individual worksheets for each unique store
store_column_index = headers.index('Store')
stores = set()
for row in ws_master.iter_rows(min_row=7, max_row=ws_master.max_row):
    store_value = row[store_column_index].value
    if store_value:
        stores.add(store_value)
existing_titles = set(wb.sheetnames)  # track titles already present in the workbook
for store in stores:
    safe_title = sanitize_sheet_title(store, existing_titles)
    ws_store = wb.create_sheet(title=safe_title)

    # Add logo
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.anchor = "B2"
        ws_store.add_image(img)

    # Merge and format E2:G2
    ws_store.merge_cells("E2:G2")
    ws_store["E2"].value = f"Inventory {prev_month_name} {prev_month_day} {current_year}"
    ws_store["E2"].alignment = Alignment(horizontal="center", vertical="center")

    # Merge and format E3:G3
    ws_store.merge_cells("E3:G3")
    ws_store["E3"].value = "SAP 4021445"
    ws_store["E3"].alignment = Alignment(horizontal="center", vertical="center")

    # Header row
    for col_num, header in enumerate(headers, start=1):
        cell = ws_store.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws_store.row_dimensions[header_row].height = 39

    # Freeze panes
    ws_store.freeze_panes = "A7"

    # Copy rows from master sheet
    for row in ws_master.iter_rows(min_row=7, max_row=ws_master.max_row):
        if row[store_column_index].value == store:
            new_row = [cell.value for cell in row]
            ws_store.append(new_row)

    # Apply borders and formatting
    for row in ws_store.iter_rows(min_row=6, max_row=ws_store.max_row, min_col=1, max_col=20):
        for cell in row:
            cell.border = thin_border

    for row in ws_store.iter_rows(min_row=7, max_row=ws_store.max_row):
        for col_name in date_columns:
            if col_name in headers:
                col_index = headers.index(col_name)
                cell = row[col_index]
                cell.number_format = "MM-DD-YY"
        for col_name in currency_columns:
            if col_name in headers:
                col_index = headers.index(col_name)
                cell = row[col_index]
                cell.number_format = '"$"#,##0.00'
                if str(cell.value).strip().lower() in ['nan', 'none']:
                    cell.value = None

    # Adjust formulas to start from row 7
    header = [cell.value for cell in ws_store[header_row]]
    col_indices = {col: idx + 1 for idx, col in enumerate(header)}
    max_row = ws_store.max_row
    for row in range(header_row + 1, max_row + 1):
        try:
            start_col = get_column_letter(col_indices['Storage Starts'])
            end_col = get_column_letter(col_indices['Storage Ends'])
            days_col = get_column_letter(col_indices['# Days In Storage'])
            charge_col = get_column_letter(col_indices['Storage Charge'])
            sqft_col = get_column_letter(col_indices['Square Footage'])
            extended_col = get_column_letter(col_indices['Extended Price'])
            ws_store[f"{days_col}{row}"] = f"=({end_col}{row}-{start_col}{row})+1"
            ws_store[f"{extended_col}{row}"] = f"=((( {charge_col}{row} * {sqft_col}{row} ) / 30) * {days_col}{row})+60"
        except KeyError:
            continue                

# Save the updated workbook
wb.save(latest_file)
print(f"Updated Excel file saved: {latest_file}")
