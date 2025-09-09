"""
Loblaw Inventory Consolidation Tool

This script processes Excel files from multiple Loblaw folders, consolidates inventory data,
and generates a master report with individual store sheets.

Author: Teams Transport
Date: 2024
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter
import os
import re
import calendar
from datetime import datetime
from typing import List, Optional, Dict, Any


class LoblawInventoryProcessor:
    """Main class for processing Loblaw inventory files."""
    
    # Define column order for consistent output
    COLUMN_ORDER = [
        'Destination Store', 'Attention', 'PO FROM LOBLAW #',
        'Shipper Order #', 'Line Up #', 'Case #', 'Case Model #', 'Serial #',
        'Estimated Ship Date', 'Arrived @ Warehouse', 'Storage Starts',
        'Storage Ends', 'Scheduled Date', 'Scheduled Time', 'Warehouse Location',
        'Trailer or Warehouse', 'Original order #', 'Original Trailer #',
        'Touched / not Touched', 'Date Stripped', 'Damage Y / N',
        'Delivery Order #', 'Delivery Trailer #', '# Days In Storage',
        'Square Footage of Case', 'Storage Charge', 'Extended Price',
        'Department', 'LH Gable', 'RH Gable', 'No Gable'
    ]
    
    # Date columns that need special formatting
    DATE_COLUMNS = [
        'Estimated Ship Date', 'Arrived @ Warehouse', 'Storage Starts',
        'Storage Ends', 'Scheduled Date', 'Date Stripped'
    ]
    
    # Valid locations for filtering
    VALID_LOCATIONS = ['trailer', 'warehouse', 'transferred']
    
    def __init__(self, output_dir: Optional[str] = None):
        """Initialize the processor with optional output directory."""
        self.output_dir = output_dir or os.path.dirname(os.path.abspath(__file__))
        self.processed_files = 0
        self.skipped_files = 0
        
    def process_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        Process a single Excel file and extract inventory data.
        
        Args:
            file_path: Path to the Excel file to process
            
        Returns:
            DataFrame with processed inventory data or None if processing failed
        """
        try:
            # Skip consolidated files
            if 'consolidated' in file_path.lower():
                print(f"  Skipping consolidated file: {os.path.basename(file_path)}")
                return None

            # Load workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            # Validate required sheets exist
            if not self._validate_sheets(wb, file_path):
                wb.close()
                return None
                
            # Extract header information
            header_info = self._extract_header_info(wb, file_path)
            if not header_info:
                wb.close()
                return None
                
            # Extract inventory data
            df = self._extract_inventory_data(wb, file_path)
            if df is None or df.empty:
                wb.close()
                return None
                
            # Apply data transformations and filters
            df = self._transform_data(df, wb, file_path)
            
            # Add header information to each row
            for key, value in header_info.items():
                df[key] = value
                
            wb.close()
            return df
            
        except Exception as e:
            print(f"  Error processing {os.path.basename(file_path)}: {str(e)}")
            return None
    
    def _validate_sheets(self, wb: Any, file_path: str) -> bool:
        """Validate that required sheets exist in the workbook."""
        required_sheets = ['Cost Estimate', 'Inventory']
        
        if not all(sheet in wb.sheetnames for sheet in required_sheets):
            print(f"  Missing required sheets in: {os.path.basename(file_path)}")
            return False
        return True
    
    def _extract_header_info(self, wb: Any, file_path: str) -> Optional[Dict[str, Any]]:
        """Extract header information from the Cost Estimate sheet."""
        try:
            ws_cost = wb['Cost Estimate']
            po_number = ws_cost['I5'].value
            
            # Skip if no PO number
            if not po_number:
                print(f"  Skipping file (no PO FROM LOBLAW # found): {os.path.basename(file_path)}")
                return None
                
            store_name = ws_cost['I10'].value
            city = ws_cost['I12'].value
            full_store_name = f"{store_name} {city}" if store_name and city else store_name or city

            return {
                'Destination Store': full_store_name,
                'Attention': ws_cost['C13'].value,
                'PO FROM LOBLAW #': po_number
            }
        except Exception as e:
            print(f"  Error extracting header info from {os.path.basename(file_path)}: {str(e)}")
            return None
    
    def _extract_inventory_data(self, wb: Any, file_path: str) -> Optional[pd.DataFrame]:
        """Extract inventory data from the Inventory sheet."""
        try:
            ws_inv = wb['Inventory']
            data_start_row = 9
            
            # Extract data rows
            data_rows = []
            for row in ws_inv.iter_rows(min_row=data_start_row, values_only=True):
                if all(cell is None for cell in row[:5]):  # Check first 5 columns
                    break
                data_rows.append(row)
            
            if not data_rows:
                print(f"  Skipping file (no inventory data found): {os.path.basename(file_path)}")
                return None
            
            # Process headers
            clean_headers = self._process_headers(ws_inv)
            
            # Create DataFrame
            df = pd.DataFrame(data_rows, columns=clean_headers)
            
            # Identify and fix time columns
            self._fix_time_columns(df, file_path)
            
            return df
            
        except Exception as e:
            print(f"  Error extracting inventory data from {os.path.basename(file_path)}: {str(e)}")
            return None
    
    def _process_headers(self, ws_inv: Any) -> List[str]:
        """Process and clean column headers from the worksheet."""
        # Get headers and subheaders
        header_row = list(ws_inv.iter_rows(min_row=7, max_row=7, values_only=True))[0]
        subheader_row = list(ws_inv.iter_rows(min_row=8, max_row=8, values_only=True))[0]
        
        clean_headers = []
        seen_headers = {}
        prev_main_header = None

        for i, (header, subheader) in enumerate(zip(header_row, subheader_row)):
            # Handle None headers (merged cells)
            if header is None:
                subheader_str = str(subheader).strip().lower() if subheader else ""
                if prev_main_header == 'Scheduled APPT' and any(t in subheader_str for t in ['time', 'tme', 'tim']):
                    header = 'Scheduled Time'
                else:
                    header = f"Unnamed_{i}"
            else:
                header = str(header).strip()
                prev_main_header = header
            
            # Handle Scheduled APPT special case
            if header == 'Scheduled APPT':
                subheader_str = str(subheader).strip().lower() if subheader else ""
                if 'date' in subheader_str:
                    header = 'Scheduled Date'
                elif any(t in subheader_str for t in ['time', 'tme', 'tim']):
                    header = 'Scheduled Time'
            
            # Handle duplicate headers
            if header in seen_headers:
                seen_headers[header] += 1
                header = f"{header}_{seen_headers[header]}"
            else:
                seen_headers[header] = 0
            
            clean_headers.append(header)
        
        return clean_headers
    
    def _fix_time_columns(self, df: pd.DataFrame, file_path: str) -> None:
        """Identify and fix time columns that might have been missed."""
        time_pattern = r'^\d{1,2}:\d{2}(?::\d{2})?(?:\s?[AP]M)?$'
        
        for col in df.columns:
            if 'time' in col.lower() and col != 'Scheduled Time':
                if df[col].astype(str).str.match(time_pattern).any():
                    df.rename(columns={col: 'Scheduled Time'}, inplace=True)
                    print(f"  Identified time column: {col} -> Scheduled Time")
        
        # Validate we have both date and time columns
        if 'Scheduled Date' in df.columns and 'Scheduled Time' not in df.columns:
            print(f"  Warning: Found Scheduled Date but missing Scheduled Time in {os.path.basename(file_path)}")
            self._find_adjacent_time_column(df)
    
    def _find_adjacent_time_column(self, df: pd.DataFrame) -> None:
        """Find time column adjacent to date column."""
        time_pattern = r'^\d{1,2}:\d{2}(?::\d{2})?(?:\s?[AP]M)?$'
        
        try:
            date_col_idx = list(df.columns).index('Scheduled Date')
            if date_col_idx + 1 < len(df.columns):
                next_col = df.columns[date_col_idx + 1]
                if df[next_col].astype(str).str.match(time_pattern).any():
                    df.rename(columns={next_col: 'Scheduled Time'}, inplace=True)
                    print(f"  Identified adjacent time column: {next_col} -> Scheduled Time")
        except (ValueError, IndexError):
            pass
    
    def _transform_data(self, df: pd.DataFrame, wb: Any, file_path: str) -> pd.DataFrame:
        """Apply all data transformations and filters."""
        # Convert date columns
        df = self._convert_date_columns(df, wb, file_path)
        
        # Apply date-based filters
        df = self._apply_date_filters(df, file_path)
        
        # Apply location filters
        df = self._apply_location_filters(df, file_path)
        
        # Adjust 'Storage Starts' based on 'Arrived @ Warehouse'
        today = datetime.today()
        first_day_this_month = datetime(today.year, today.month, 1)
        last_day_prev_month = first_day_this_month - pd.Timedelta(days=1)
        first_day_prev_month = datetime(last_day_prev_month.year, last_day_prev_month.month, 1)

        if 'Arrived @ Warehouse' in df.columns and 'Storage Starts' in df.columns:
            df['Storage Starts'] = df.apply(
                lambda row: row['Arrived @ Warehouse']
                if pd.notnull(row['Arrived @ Warehouse']) and first_day_prev_month <= row['Arrived @ Warehouse'] <= last_day_prev_month
                else first_day_prev_month,
                axis=1
            )

        # Adjust 'Storage Ends' based on 'Scheduled Date'
        if 'Scheduled Date' in df.columns and 'Storage Ends' in df.columns:
            df['Storage Ends'] = df.apply(
                lambda row: row['Scheduled Date']
                if pd.notnull(row['Scheduled Date']) and first_day_prev_month <= row['Scheduled Date'] <= last_day_prev_month
                else last_day_prev_month,
                axis=1
            )
        return df
    
    def _convert_date_columns(self, df: pd.DataFrame, wb: Any, file_path: str) -> pd.DataFrame:
        """Convert date columns to proper datetime format."""
        ws_inv = wb['Inventory']
        data_start_row = 9
        
        # Standard date columns
        standard_date_columns = ['Storage Starts', 'Storage Ends', 'Estimated Ship Date', 'Arrived @ Warehouse']
        
        for col in standard_date_columns:
            if col in df.columns:
                # Try direct conversion first
                df[col] = pd.to_datetime(df[col], format='%Y-%m-%d', errors='coerce')
                
                # Handle formula cells if needed
                if df[col].isna().any():
                    self._handle_formula_dates(df, col, ws_inv, data_start_row)
        
        # Special handling for 'Arrived @ Warehouse'
        if 'Arrived @ Warehouse' in df.columns:
            df['Arrived @ Warehouse'] = pd.to_datetime(df['Arrived @ Warehouse'], format='%m-%d-%y', errors='coerce')
        
        # Special handling for 'Storage Ends'
        if 'Storage Ends' in df.columns:
            df['Storage Ends'] = pd.to_datetime(df['Storage Ends'], format='%m-%d-%y', errors='coerce')
        
        return df

    def _handle_formula_dates(self, df: pd.DataFrame, col: str, ws_inv: Any, data_start_row: int) -> None:
        """Handle date columns that contain formulas."""
        try:
            col_idx = df.columns.get_loc(col)
            for i in range(len(df)):
                if pd.isna(df.at[i, col]):
                    cell_value = ws_inv.cell(row=i + data_start_row, column=col_idx + 1).value
                    df.at[i, col] = pd.to_datetime(cell_value, errors='coerce')
        except Exception as e:
            print(f"  Warning: Could not evaluate {col} formulas: {str(e)}")
    
    def _apply_date_filters(self, df: pd.DataFrame, file_path: str) -> pd.DataFrame:
        """Apply date-based filters to the data."""
        initial_count = len(df)
        
        # Filter 'Arrived @ Warehouse' dates
        if 'Arrived @ Warehouse' in df.columns:
            lower_cutoff = pd.to_datetime('2000-01-01')
            upper_cutoff = pd.to_datetime(datetime.now().strftime('%Y-%m-01'))
            
            df = df[(df['Arrived @ Warehouse'] > lower_cutoff) & 
                   (df['Arrived @ Warehouse'] < upper_cutoff)]
            
            if len(df) != initial_count:
                print(f"  Filtered out {initial_count - len(df)} rows based on 'Arrived @ Warehouse' date condition")
                initial_count = len(df)
        
        # Filter 'Storage Ends' dates
        if 'Storage Ends' in df.columns:
            today = datetime.now()
            first_of_prev_month = (datetime(today.year, today.month - 1, 1) 
                                 if today.month > 1 
                                 else datetime(today.year - 1, 12, 1))
            early_cutoff = pd.to_datetime('2000-01-31')
            
            df = df[
                (df['Storage Ends'].isna()) |
                (df['Storage Ends'] <= early_cutoff) |
                (df['Storage Ends'] >= first_of_prev_month)
            ]
            
            if len(df) != initial_count:
                print(f"  Filtered out {initial_count - len(df)} rows based on 'Storage Ends' date condition")
        
        return df
    
    def _apply_location_filters(self, df: pd.DataFrame, file_path: str) -> pd.DataFrame:
        """Apply location-based filters to the data."""
        if 'Trailer or Warehouse' not in df.columns:
            return df
            
        initial_count = len(df)
        df = df[df['Trailer or Warehouse'].str.lower().isin(self.VALID_LOCATIONS)]
        
        if len(df) != initial_count:
            print(f"  Filtered out {initial_count - len(df)} rows based on 'Trailer or Warehouse' value condition")
        
        return df
    
    def format_date_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Format datetime columns for display."""
        # Format date columns to display only date (no time)
        for col in self.DATE_COLUMNS:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
                df[col] = df[col].dt.date
        
        # Handle Scheduled Time formatting
        if 'Scheduled Time' in df.columns:
            df['Scheduled Time'] = df['Scheduled Time'].astype(str)
            # Clean up time formatting
            df['Scheduled Time'] = df['Scheduled Time'].str.replace(r'\.\d+$', '', regex=True)
            # Convert HH:MM:SS to HH:MM
            df['Scheduled Time'] = df['Scheduled Time'].apply(
                lambda x: x[:5] if len(x) >= 8 and x.count(':') >= 2 else x
            )
            # Standardize AM/PM formatting
            df['Scheduled Time'] = df['Scheduled Time'].str.upper().str.replace(' ', '')
        
        return df
    
    def process_folders(self, folders: List[str]) -> pd.DataFrame:
        """
        Process all Excel files in the specified folders.
        
        Args:
            folders: List of folder paths to process
            
        Returns:
            Consolidated DataFrame with all processed data
        """
        all_data_frames = []
        self.processed_files = 0
        self.skipped_files = 0
        
        for folder_path in folders:
            if not os.path.exists(folder_path):
                print(f"\nFolder not found: {folder_path}")
                continue
                
            print(f"\nProcessing folder: {folder_path}")
            
            for file in os.listdir(folder_path):
                if file.endswith('.xlsx'):
                    full_path = os.path.join(folder_path, file)
                    df = self.process_file(full_path)
                    if df is not None and not df.empty:
                        all_data_frames.append(df)
                        self.processed_files += 1
                    else:
                        self.skipped_files += 1
        
        if not all_data_frames:
            raise ValueError("No data was processed. Check your files and try again.")
        
        # Consolidate all data
        return self._consolidate_dataframes(all_data_frames)
    
    def _consolidate_dataframes(self, data_frames: List[pd.DataFrame]) -> pd.DataFrame:
        """Consolidate multiple DataFrames into a single DataFrame."""
        # Get all possible columns
        all_columns = set()
        for df in data_frames:
            all_columns.update(df.columns)
        
        # Reindex each DataFrame with all possible columns
        dfs_reindexed = []
        for df in data_frames:
            missing_cols = all_columns - set(df.columns)
            for col in missing_cols:
                df[col] = pd.NA
            # Drop completely empty columns
            df = df.dropna(axis=1, how='all')
            dfs_reindexed.append(df)
        
        # Concatenate all DataFrames
        consolidated_df = pd.concat(dfs_reindexed, ignore_index=True)
        
        # Clean up column names
        consolidated_df.columns = [col.split('_')[0] if '_' in col else col 
                                  for col in consolidated_df.columns]
        
        # Format date columns
        consolidated_df = self.format_date_columns(consolidated_df)
        
        # Reorder columns
        columns_order = [col for col in self.COLUMN_ORDER if col in consolidated_df.columns]
        consolidated_df = consolidated_df[columns_order]
        
        return consolidated_df
    
    def create_individual_store_sheets(self, df: pd.DataFrame, wb: Any) -> None:
        """Create individual sheets for each store."""
        # Get previous month information
        today = datetime.today()
        first_day_this_month = datetime(today.year, today.month, 1)
        last_day_prev_month = first_day_this_month - pd.Timedelta(days=1)
        prev_month_name = last_day_prev_month.strftime("%B")
        prev_month_last_day = last_day_prev_month.day
        current_year = today.year
        sheet_suffix = last_day_prev_month.strftime("%m.%y")
        
        # Create sheets for each unique store
        for store in df['Destination Store'].dropna().unique():
            store_df = df[df['Destination Store'] == store]
            
            # Generate sheet name
            simple_store = self._extract_simple_store(store)
            sheet_name = f"{simple_store} {sheet_suffix}"
            sheet_name = re.sub(r"[\\/*?:\[\]]", "", sheet_name)[:31]
            
            # Create new sheet
            ws = wb.create_sheet(title=sheet_name)
            
            # Add header information
            ws.append([f"Inventory {prev_month_name} {prev_month_last_day} {current_year}"])
            ws.append(["SAP 4021445"])
            ws.append([])  # Spacer row
            
            # Add column headers
            ws.append(self.COLUMN_ORDER)
            
            # Format header row
            for cell in ws[4]:
                cell.font = Font(bold=True)
            
            # Add inventory data
            for _, row in store_df.iterrows():
                ws.append([row.get(col, "") for col in self.COLUMN_ORDER])
    
    def _extract_simple_store(self, store_name: str) -> str:
        """Extract simple store name from full store name."""
        match = re.search(r'#\d+', store_name)
        if match:
            start_index = match.start()
            after_hash = store_name[start_index:]
            space_match = re.search(r'\s', after_hash)
            if space_match:
                end_index = start_index + space_match.start()
                return store_name[:end_index].strip()
        return store_name.strip()
    
    def save_consolidated_report(self, df: pd.DataFrame, filename: str) -> str:
        # Compute previous month boundaries
        today = datetime.today()
        first_day_this_month = datetime(today.year, today.month, 1)
        last_day_prev_month = first_day_this_month - pd.Timedelta(days=1)
        first_day_prev_month = datetime(last_day_prev_month.year, last_day_prev_month.month, 1)

        # Convert relevant columns to datetime
        df['Arrived @ Warehouse'] = pd.to_datetime(df['Arrived @ Warehouse'], errors='coerce')
        df['Scheduled Date'] = pd.to_datetime(df['Scheduled Date'], errors='coerce')

        # Apply logic for 'Storage Starts'
        df['Storage Starts'] = df['Arrived @ Warehouse'].apply(
            lambda x: x if pd.notnull(x) and first_day_prev_month <= x <= last_day_prev_month else first_day_prev_month
        )

        # Apply logic for 'Storage Ends'
        df['Storage Ends'] = df['Scheduled Date'].apply(
            lambda x: x if pd.notnull(x) and first_day_prev_month <= x <= last_day_prev_month else last_day_prev_month
        )

        """
        Save the consolidated report with individual store sheets.
        
        Args:
            df: Consolidated DataFrame
            filename: Output filename
            
        Returns:
            Full path to the saved file
        """
        output_path = os.path.join(self.output_dir, filename)
        
        # Drop specified columns before saving
        columns_to_hide = ['Estimated Ship Date', 'Trailer or Warehouse', 'Original order #', 'Original Trailer #', 'Touched / not Touched', 'Date Stripped', 'Delivery Order #', 'Department', 'LH Gable', 'RH Gable', 'No Gable']
        df = df.drop(columns=[col for col in columns_to_hide if col in df.columns], errors='ignore')        

        # Save main data to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl', datetime_format='MM-DD-YY') as writer:
            df.to_excel(writer, index=False)
        
        # Load workbook and add individual store sheets
        wb = openpyxl.load_workbook(output_path)
        wb.active.title = 'Master'
        
        # Inject formulas into the master sheet
        ws = wb['Master']
        header = [cell.value for cell in ws[1]]
        col_indices = {col: idx + 1 for idx, col in enumerate(header)}

        for row in range(2, ws.max_row + 1):
            start_col = get_column_letter(col_indices['Storage Starts'])
            end_col = get_column_letter(col_indices['Storage Ends'])
            days_col = get_column_letter(col_indices['# Days In Storage'])
            charge_col = get_column_letter(col_indices['Storage Charge'])
            sqft_col = get_column_letter(col_indices['Square Footage of Case'])
            extended_col = get_column_letter(col_indices['Extended Price'])

            ws[f"{days_col}{row}"] = f"=({end_col}{row}-{start_col}{row})+1"
            ws[f"{extended_col}{row}"] = f"=((({charge_col}{row}*{sqft_col}{row})/30)*{days_col}{row})+60"

        # Save updated workbook
        wb.save(output_path)
        
        return output_path


def main():
    """Main function to run the inventory consolidation process."""
    # Configuration - Update these paths as needed
    folders_to_process = [
        r'C:\Users\cgatz\TEAMS Transport\TEAMS - Johanne-Work\Loblaw\Alberta',
        r'C:\Users\cgatz\TEAMS Transport\TEAMS - Johanne-Work\Loblaw\BC',
        r'C:\Users\cgatz\TEAMS Transport\TEAMS - Johanne-Work\Loblaw\Manitoba',
        r'C:\Users\cgatz\TEAMS Transport\TEAMS - Johanne-Work\Loblaw\NL, NS, PE, NB',
        r'C:\Users\cgatz\TEAMS Transport\TEAMS - Johanne-Work\Loblaw\NWT, YT',
        r'C:\Users\cgatz\TEAMS Transport\TEAMS - Johanne-Work\Loblaw\Ontario',
        r'C:\Users\cgatz\TEAMS Transport\TEAMS - Johanne-Work\Loblaw\Quebec',
        r'C:\Users\cgatz\TEAMS Transport\TEAMS - Johanne-Work\Loblaw\Saskatchewan',
    ]
    
    try:
        # Initialize processor
        processor = LoblawInventoryProcessor()
        
        # Process all folders
        print("Starting Loblaw Inventory Consolidation Process...")
        consolidated_data = processor.process_folders(folders_to_process)
        
        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"LoblawsConsolidatedInventory_{timestamp}.xlsx"
        
        # Save consolidated report
        output_path = processor.save_consolidated_report(consolidated_data, output_filename)
        
        # Print summary
        print(f"\nProcessing complete:")
        print(f"- Processed folders: {len(folders_to_process)}")
        print(f"- Successfully processed files: {processor.processed_files}")
        print(f"- Skipped files: {processor.skipped_files}")
        print(f"\nSaved consolidated data to: {output_path}")
        print(f"\nSummary:")
        print(f"Total rows: {len(consolidated_data)}")
        print(f"Columns: {len(consolidated_data.columns)}")
        
        # Open output directory
        try:
            os.startfile(processor.output_dir)
        except:
            print(f"\nOutput directory: {processor.output_dir}")
            
    except Exception as e:
        print(f"\nError: {str(e)}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())