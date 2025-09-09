import pandas as pd
from openpyxl import load_workbook
from fpdf import FPDF
import os

# Load the workbook
workbook_path = "LoblawsConsolidatedInventory_20250811_152210.xlsx"
wb = load_workbook(workbook_path, data_only=True)

# Create output directory for PDFs
output_dir = "sheet_pdfs"
os.makedirs(output_dir, exist_ok=True)

# Iterate through all sheets except 'master'
for sheet_name in wb.sheetnames:
    if sheet_name.lower() == 'master':
        continue

    # Load sheet into a DataFrame
    df = pd.read_excel(workbook_path, sheet_name=sheet_name, engine='openpyxl')

    # Create a PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)

    # Add sheet name as title
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(200, 10, txt=sheet_name, ln=True, align='C')
    pdf.set_font("Arial", size=10)

    # Add table content
    col_widths = [pdf.get_string_width(str(col)) + 2 for col in df.columns]
    row_height = 8

    # Header
    for i, col in enumerate(df.columns):
        pdf.cell(col_widths[i], row_height, str(col), border=1)
    pdf.ln(row_height)

    # Rows
    for _, row in df.iterrows():
        for i, item in enumerate(row):
            pdf.cell(col_widths[i], row_height, str(item), border=1)
        pdf.ln(row_height)

    # Save PDF
    pdf_output_path = os.path.join(output_dir, f"{sheet_name}.pdf")
    pdf.output(pdf_output_path)

print(f"PDFs for all sheets except 'master' have been saved in the '{output_dir}' directory.")
